from typing import List

from openpyxl import Workbook

FIRST_QUESTION_COLUMN_NUMBER = 1
FIRST_QUESTION_ROW_NUMBER = 4


class Answer:
    text: str
    is_right: bool

    def __init__(self, text: str, is_right: bool):
        self.text = text
        self.is_right = is_right

    def __str__(self) -> str:
        return self.text


class Question:
    text: str
    answers: List[Answer]

    def __init__(self, text: str, answers: List[Answer] = None):
        self.text = text
        self.answers = answers

    def __str__(self) -> str:
        return self.text


class Test:
    name: str
    questions: List[Question]

    def __init__(self, name: str, questions: List[Question] = None):
        self.name = name
        self.questions = questions

    def __str__(self) -> str:
        return self.name


def parse_tests(excel_file: Workbook) -> List[Test]:
    tests = []

    for sheet_name in excel_file.sheetnames:
        sheet = excel_file[sheet_name]

        if not sheet["A1"].value:
            raise ValueError(f"Test name on sheet {sheet_name} is empty")

        test = Test(name=sheet["A1"].value)
        questions_dict = {}

        row_number = FIRST_QUESTION_ROW_NUMBER

        while (question_number_cell := sheet.cell(column=FIRST_QUESTION_COLUMN_NUMBER,
                                                  row=row_number)) and question_number_cell.value:
            question_number = int(question_number_cell.value)
            question_text_cell = sheet.cell(column=FIRST_QUESTION_COLUMN_NUMBER + 1, row=row_number)
            answer_text_cell = sheet.cell(column=FIRST_QUESTION_COLUMN_NUMBER + 2, row=row_number)
            is_right_answer_cell = sheet.cell(column=FIRST_QUESTION_COLUMN_NUMBER + 3, row=row_number)

            if question_number not in questions_dict:
                if not question_text_cell.value:
                    raise ValueError(f"Question #{question_number} on sheet {sheet_name} is empty")
                questions_dict[question_number] = Question(text=str(question_text_cell.value), answers=[])

            question: Question = questions_dict[question_number]

            if not answer_text_cell.value:
                raise ValueError(f"Answer for question #{question_number} on sheet {sheet_name} is empty")

            question.answers.append(Answer(text=str(answer_text_cell.value), is_right=bool(is_right_answer_cell.value)))

            row_number += 1

        test.questions = [question for question in questions_dict.values()]
        tests.append(test)

    return tests
