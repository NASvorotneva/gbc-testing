from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles.colors import COLOR_INDEX, Color
from openpyxl.styles.fills import PatternFill

from utils.db_api.database import Test, Result, User, Question, UserAnswer, Answer


RED_COLOUR = Color(rgb=COLOR_INDEX[2])
RED_FILL = PatternFill(patternType='solid', fgColor=RED_COLOUR)

GREEN_COLOUR = Color(rgb=COLOR_INDEX[3])
GREEN_FILL = PatternFill(patternType='solid', fgColor=GREEN_COLOUR)


async def export_test_results(test: Test) -> BytesIO:
    users_results = await Result.filter(Result.test_id == test.id)
    questions = await Question.filter(Question.test_id == test.id)
    wb = Workbook()
    active_sheet = wb.active
    active_sheet.title = "Результаты"

    active_sheet.cell(column=1, row=1, value=test.name)
    active_sheet.cell(column=1, row=3, value="ID Телеграм")
    active_sheet.cell(column=2, row=3, value="Полное имя")
    active_sheet.cell(column=3, row=3, value="Username")
    active_sheet.cell(column=4, row=3, value="Номер телефона")

    x = 5

    for x_index, question in enumerate(questions):
        active_sheet.cell(column=x + x_index, row=3, value=question.text)

    active_sheet.cell(column=x + len(questions), row=3, value="Итого")

    y = 4

    for y_index, user_result in enumerate(users_results):
        user = await User.get(User.id == user_result.user_id)

        active_sheet.cell(column=1, row=y + y_index, value=str(user.id))
        active_sheet.cell(column=2, row=y + y_index, value=user.full_name)
        active_sheet.cell(column=3, row=y + y_index, value=user.username)
        active_sheet.cell(column=4, row=y + y_index, value=user.phone_number)

        for x_index, question in enumerate(questions):
            user_answer = await UserAnswer.get(UserAnswer.question_id == question.id, UserAnswer.user_id == user.id)
            answer = await Answer.get(Answer.id == user_answer.answer_id)
            answer_cell = active_sheet.cell(column=x + x_index, row=y + y_index, value=answer.text)
            answer_cell.fill = GREEN_FILL if answer.is_right else RED_FILL

        active_sheet.cell(column=x + len(questions), row=y + y_index, value=f"{user_result.result}%")

    bytes_io = BytesIO()
    wb.save(bytes_io)
    bytes_io.seek(0)

    return bytes_io

